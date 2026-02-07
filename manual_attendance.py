import tkinter as tk
from datetime import date
import os
from openpyxl import Workbook, load_workbook

# ---------- CONFIG ----------
ATT_FILE = "attendance.xlsx"
REPORT_FILE = "daily_report.xlsx"

EMPLOYEES = {
    "E101": {"name": "Rahul", "pin": "1234"},
    "E102": {"name": "Aman", "pin": "2345"},
    "E103": {"name": "Neha", "pin": "3456"},
    "E104": {"name": "Priya", "pin": "4567"},
}

ADMIN_PASSWORD = "admin123"

# ---------- EXCEL SETUP ----------
def init_excel():
    if not os.path.exists(ATT_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee ID", "Name", "Date", "Slot", "Status"])
        wb.save(ATT_FILE)

def already_marked(emp_id, today):
    wb = load_workbook(ATT_FILE)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] == emp_id and row[2] == today:
            return True
    return False

# ---------- ATTENDANCE ----------
def mark_attendance(emp_id):
    today = date.today().strftime("%d-%m-%Y")
    slot = slot_var.get()

    if slot not in ("Morning", "Lunch"):
        msg.config(text="Select Morning or Lunch slot", fg="red")
        return

    if already_marked(emp_id, today):
        msg.config(text="Attendance already marked today", fg="orange")
        return

    status = "Present" if slot == "Morning" else "Half Day"

    try:
        wb = load_workbook(ATT_FILE)
        ws = wb.active
        ws.append([
            emp_id,
            EMPLOYEES[emp_id]["name"],
            today,
            slot,
            status
        ])
        wb.save(ATT_FILE)
        msg.config(text=f"{EMPLOYEES[emp_id]['name']} marked {status}", fg="green")
    except PermissionError:
        msg.config(text="Close attendance.xlsx and try again", fg="red")

# ---------- EMPLOYEE LOGIN ----------
def employee_login():
    emp_id = id_entry.get().strip()
    pin = pin_entry.get().strip()

    if emp_id in EMPLOYEES and EMPLOYEES[emp_id]["pin"] == pin:
        mark_attendance(emp_id)
    else:
        msg.config(text="Invalid ID or PIN", fg="red")

# ---------- ADMIN ----------
def mark_absent():
    today = date.today().strftime("%d-%m-%Y")

    try:
        wb = load_workbook(ATT_FILE)
        ws = wb.active

        present_ids = set()
        for row in ws.iter_rows(values_only=True):
            if row[2] == today:
                present_ids.add(row[0])

        for emp_id, data in EMPLOYEES.items():
            if emp_id not in present_ids:
                ws.append([
                    emp_id,
                    data["name"],
                    today,
                    "None",
                    "Absent"
                ])

        wb.save(ATT_FILE)
        admin_msg.config(text="Absentees marked", fg="green")
    except PermissionError:
        admin_msg.config(text="Close attendance.xlsx first", fg="red")

def generate_report():
    today = date.today().strftime("%d-%m-%Y")

    try:
        wb = load_workbook(ATT_FILE)
        ws = wb.active

        report = Workbook()
        rws = report.active
        rws.append(["Employee ID", "Name", "Date", "Slot", "Status"])

        for row in ws.iter_rows(values_only=True):
            if row[2] == today:
                rws.append(row)

        report.save(REPORT_FILE)
        admin_msg.config(text="Daily report generated", fg="green")
    except PermissionError:
        admin_msg.config(text="Close daily_report.xlsx first", fg="red")

def admin_login():
    if admin_entry.get() != ADMIN_PASSWORD:
        admin_status.config(text="Wrong admin password", fg="red")
        return

    win = tk.Toplevel(root)
    win.title("Admin Panel")
    win.geometry("300x250")

    tk.Label(win, text="Admin Panel", font=("Arial", 14, "bold")).pack(pady=10)

    tk.Button(win, text="Mark Absentees", width=22, command=mark_absent).pack(pady=10)
    tk.Button(win, text="Generate Daily Report", width=22, command=generate_report).pack(pady=10)

    global admin_msg
    admin_msg = tk.Label(win, text="")
    admin_msg.pack(pady=10)

# ---------- GUI ----------
init_excel()

root = tk.Tk()
root.title("Manual Attendance System")
root.geometry("420x560")

tk.Label(root, text="Manual Attendance System", font=("Arial", 16, "bold")).pack(pady=10)

tk.Label(root, text="Employee ID").pack()
id_entry = tk.Entry(root)
id_entry.pack()

tk.Label(root, text="PIN").pack()
pin_entry = tk.Entry(root, show="*")
pin_entry.pack()

tk.Label(root, text="Select Slot").pack(pady=5)
slot_var = tk.StringVar()

tk.Radiobutton(root, text="Morning (Full Day)", variable=slot_var, value="Morning").pack()
tk.Radiobutton(root, text="Lunch (Half Day)", variable=slot_var, value="Lunch").pack()

tk.Button(root, text="Mark Attendance", command=employee_login).pack(pady=10)

msg = tk.Label(root, text="", font=("Arial", 11))
msg.pack(pady=10)

# ---- Admin Login ----
tk.Label(root, text="Admin Login", font=("Arial", 12, "bold")).pack(pady=15)

admin_entry = tk.Entry(root, show="*")
admin_entry.pack()

tk.Button(root, text="Admin Login", command=admin_login).pack(pady=5)

admin_status = tk.Label(root, text="")
admin_status.pack()

root.mainloop()
