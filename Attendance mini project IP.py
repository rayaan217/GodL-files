import tkinter as tk
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

# ---------------- CONFIG ---------------- #

FILE_NAME = "attendance.xlsx"
REPORT_FILE = "daily_report.xlsx"

EMPLOYEES = {
    "E101": {"name": "Rahul", "pin": "1234"},
    "E102": {"name": "Aman", "pin": "2345"},
    "E103": {"name": "Neha", "pin": "3456"},
    "E104": {"name": "Priya", "pin": "4567"},
}

ADMIN_PASSWORD = "admin123"

# ---------------- EXCEL SETUP ---------------- #

def setup_excel():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.append(["Employee ID", "Name", "Date", "Time", "Status"])
        wb.save(FILE_NAME)

def attendance_exists(emp_id, date_today):
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] == emp_id and row[2] == date_today:
            return True
    return False

# ---------------- ATTENDANCE LOGIC ---------------- #

def mark_attendance(emp_id):
    now = datetime.now()
    today = now.strftime("%d-%m-%Y")
    current_time = now.time()

    morning_start = datetime.strptime("09:00", "%H:%M").time()
    morning_end   = datetime.strptime("09:30", "%H:%M").time()

    lunch_start = datetime.strptime("13:00", "%H:%M").time()
    lunch_end   = datetime.strptime("13:30", "%H:%M").time()

    if attendance_exists(emp_id, today):
        status_label.config(text="Attendance already marked today!", fg="orange")
        return

    if morning_start <= current_time <= morning_end:
        status = "Present"
    elif lunch_start <= current_time <= lunch_end:
        status = "Half Day"
    else:
        status_label.config(text="Attendance not allowed at this time!", fg="red")
        return

    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append([
        emp_id,
        EMPLOYEES[emp_id]["name"],
        today,
        now.strftime("%H:%M:%S"),
        status
    ])
    wb.save(FILE_NAME)

    status_label.config(
        text=f"{EMPLOYEES[emp_id]['name']} marked as {status}",
        fg="green"
    )

# ---------------- EMPLOYEE LOGIN ---------------- #

def employee_login():
    emp_id = id_entry.get()
    pin = pin_entry.get()

    if emp_id in EMPLOYEES and EMPLOYEES[emp_id]["pin"] == pin:
        mark_attendance(emp_id)
    else:
        status_label.config(text="Invalid Employee ID or PIN", fg="red")

# ---------------- ADMIN FUNCTIONS ---------------- #

def mark_absentees():
    today = datetime.now().strftime("%d-%m-%Y")

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    marked = set()
    for row in ws.iter_rows(values_only=True):
        if row[2] == today:
            marked.add(row[0])

    for emp_id, data in EMPLOYEES.items():
        if emp_id not in marked:
            ws.append([
                emp_id,
                data["name"],
                today,
                "N/A",
                "Absent"
            ])

    wb.save(FILE_NAME)
    admin_status.config(text="Absentees marked successfully", fg="green")

def generate_daily_report():
    today = datetime.now().strftime("%d-%m-%Y")

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    report = Workbook()
    rws = report.active
    rws.append(["Employee ID", "Name", "Date", "Time", "Status"])

    for row in ws.iter_rows(values_only=True):
        if row[2] == today:
            rws.append(row)

    report.save(REPORT_FILE)
    admin_status.config(text="Daily report generated", fg="green")

def admin_panel():
    admin_win = tk.Toplevel(root)
    admin_win.title("Admin Panel")
    admin_win.geometry("350x300")

    tk.Label(admin_win, text="Admin Panel", font=("Arial", 14, "bold")).pack(pady=10)

    tk.Button(admin_win, text="Mark Absentees", width=25, command=mark_absentees).pack(pady=10)
    tk.Button(admin_win, text="Generate Daily Report", width=25, command=generate_daily_report).pack(pady=10)

    global admin_status
    admin_status = tk.Label(admin_win, text="")
    admin_status.pack(pady=10)

def admin_login():
    if admin_entry.get() == ADMIN_PASSWORD:
        admin_panel()
    else:
        admin_status_main.config(text="Wrong Admin Password", fg="red")

# ---------------- GUI ---------------- #

setup_excel()

root = tk.Tk()
root.title("Employee Attendance System")
root.geometry("420x520")

tk.Label(root, text="Employee Attendance System", font=("Arial", 16, "bold")).pack(pady=10)

tk.Label(root, text="Employee ID").pack()
id_entry = tk.Entry(root)
id_entry.pack()

tk.Label(root, text="PIN").pack()
pin_entry = tk.Entry(root, show="*")
pin_entry.pack()

tk.Button(root, text="Mark Attendance", command=employee_login).pack(pady=10)

status_label = tk.Label(root, text="", font=("Arial", 11))
status_label.pack(pady=10)

# -------- Admin Login -------- #

tk.Label(root, text="Admin Login", font=("Arial", 12, "bold")).pack(pady=15)

admin_entry = tk.Entry(root, show="*")
admin_entry.pack()

tk.Button(root, text="Admin Login", command=admin_login).pack(pady=5)

admin_status_main = tk.Label(root, text="")
admin_status_main.pack()

root.mainloop()
